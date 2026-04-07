"""
Enslaved Ontology Class Embedding Evaluator
============================================
Computes cosine similarity between ontology class tokens/sentences
and competency questions (CQs) using four embedding models.

Usage:
    python embed_eval.py                        # raw class names (V1)
    python embed_eval.py --to_sentence          # convert to sentences first (V2)
    python embed_eval.py --both                 # run both and save to separate sheets
    python embed_eval.py --models bert bge      # select specific models
"""

import argparse
import time
from pathlib import Path

import numpy as np
import torch
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# ──────────────────────────────────────────────────────────────────────────────
# DATA
# ──────────────────────────────────────────────────────────────────────────────

CLASSES = [
    "ParticipantRoleRecord",
    "SexRecord",
    "AgentRecord",
    "InterAgentRelationshipTypes",
    "PersonRecord",
    "TemporalExtent",
    "SexTypes",
    "ProvenanceActivity",
    "ExternalReferent",
    "ECVO",
    "InterAgentRelationshipRecord",
    "EventTypes",
    "MatchType",
    "Event",
    "ResearchProjectContributorRole",
    "SpatialExtent",
    "Match",
    "Description",
    "DocumentTypes",
    "OccupationRecord",
    "TimeSpan",
    "ExternalReference",
    "LicenseInformation",
    "NameRecord",
    "NameVariant",
    "Organization",
    "PlaceTypeCV",
    "OriginRecord",
    "Researcher",
    "Coordinates",
    "Agent",
    "Place",
    "Person",
    "ResearchProjectPIRole",
    "ResearchProject",
    "Occupations",
    "ParticipantRoleTypes",
    "EntityWithProvenance",
    "RaceRecord",
    "PersonStatusRecord",
    "AgeCategory",
    "PlaceCV",
    "AgeRecord",
    "PSCategories",
]

COMPETENCY_QUESTIONS = [
    "What are all the enslaved people in County X in the century X?",
    "What are all the enslaved individuals categorized by gender, ethnicity, and location (e.g., on plantation X or of XXXX ethnicity)?",
    "What are all the available population counts of enslaved people by place and year?",
    "What are all the documented literacy statuses of enslaved individuals?",
    "What are all the geographical places associated with enslaved people and their relationships to those places?",
    "What are all the common names of enslaved children during the Revolutionary War era?",
    "What are all the enslaved individuals and who enslaved them, including locations?",
    "Who are all the children associated with enslaved individuals?",
    "What are all the available records about individuals identified as enslaved or formerly enslaved in specific locations and periods?",
    "What are all the enslaved children recorded in specific locations and time periods?",
    "What are all the enslaved people associated with a specific owner in a specific place and time, and how many were there?",
    "What are all the enslaved individuals who contributed labor to construction projects, such as the state capitol, and what were their roles?",
    "What are all the known addresses or properties that were inhabited by enslaved individuals?",
    "What are all the enslaved individuals transported from specific ports and their demographic breakdowns over time?",
    "What are all the known slave rebellions, their participants, locations, and dates?",
    "What are all the known relatives of enslaved individuals, including their relationships and locations?",
    "What are all the median ages of enslaved individuals who escaped in location X during time period X?",
    "What are all the average ages at death of enslaved individuals?",
    "What are all the documented skill categories associated with enslaved individuals, broken down by ethnic group, and their percentages?",
    "What are all the recorded marriage events between enslaved individuals, categorized by same or different ethnic groupings, and their percentages?",
    "What are all the known economic values assigned to enslaved individuals, categorized by ethnicity, skills, gender, and period?",
    "What are all the life outcomes (e.g., sale, manumission, death) of enslaved individuals disembarked from ship XXX?",
    "What are all the enslaved individuals who lived in location X in the time period XXX?",
    "What are all the enslaved individuals who were later recorded as slave owners, including details of their ownership activities?",
]

# ──────────────────────────────────────────────────────────────────────────────
# CLASS → SENTENCE CONVERSION
# ──────────────────────────────────────────────────────────────────────────────

# Manual human-readable sentence descriptions for each class.
# These unpack CamelCase names into domain-meaningful natural language.
CLASS_SENTENCES = {
    "ParticipantRoleRecord":         "ParticipantRoleRecord - A record describing the role a participant played in a specific event.",
    "SexRecord":                     "SexRecord - A record documenting the biological sex of an individual.",
    "AgentRecord":                   "AgentRecord - A record representing an agent, which may be a person or organization.",
    "InterAgentRelationshipTypes":   "InterAgentRelationshipTypes - A controlled vocabulary of relationship types between agents, such as family, ownership, or social ties.",
    "PersonRecord":                  "PersonRecord - A record containing biographical and identifying information about a person.",
    "TemporalExtent":                "TemporalExtent - A representation of a time span or date range associated with an event or entity.",
    "SexTypes":                      "SexTypes - A controlled vocabulary of sex or gender categories used to classify individuals.",
    "ProvenanceActivity":            "ProvenanceActivity - An activity that documents the origin, history, or custody chain of a record or entity.",
    "ExternalReferent":              "ExternalReferent - A reference to an entity or record held in an external database or system.",
    # "ECVO":                          "ECVO - An enslaved community vocabulary ontology term used to categorize concepts related to enslaved people.",
    "InterAgentRelationshipRecord":  "InterAgentRelationshipRecord - A record documenting a specific relationship between two agents, including its type and temporal extent.",
    "EventTypes":                    "EventTypes - A controlled vocabulary classifying types of historical events such as sale, birth, death, or manumission.",
    "MatchType":                     "MatchType - A classification of how confidently two records have been matched to represent the same entity.",
    "Event":                         "Event - A historical occurrence involving one or more agents, associated with a time and place.",
    "ResearchProjectContributorRole":"ResearchProjectContributorRole - The role a contributor plays within a research project, such as data curator or analyst.",
    "SpatialExtent":                 "SpatialExtent - A geographical area or bounding region associated with an entity or event.",
    "Match":                         "Match - A linkage asserting that two or more records refer to the same real-world entity.",
    "Description":                   "Description - A textual or structured description providing additional context for an entity.",
    "DocumentTypes":                 "DocumentTypes - A controlled vocabulary classifying the types of historical documents, such as census, deed, or ship manifest.",
    "OccupationRecord":              "OccupationRecord - A record documenting the occupation or skilled trade of an individual.",
    "TimeSpan":                      "TimeSpan - A specific period of time defined by start and end dates, used to bound events or statuses.",
    "ExternalReference":             "ExternalReference - A link or citation pointing to a related record or resource in an external system.",
    "LicenseInformation":            "LicenseInformation - Metadata describing the licensing terms governing the use of a dataset or record.",
    "NameRecord":                    "NameRecord - A record capturing a name used to identify an individual, including variants and aliases.",
    "NameVariant":                   "NameVariant - An alternative spelling, abbreviation, or form of an individual's name.",
    "Organization":                  "Organization - An institution, group, or corporate body that acts as an agent in historical records.",
    "PlaceTypeCV":                   "PlaceTypeCV - A controlled vocabulary classifying types of places, such as plantation, county, port, or city.",
    "OriginRecord":                  "OriginRecord - A record documenting the geographic or ethnic origin of an individual.",
    "Researcher":                    "Researcher - A person who contributes to a research project by collecting, curating, or analyzing historical records.",
    "Coordinates":                   "Coordinates - Geographic latitude and longitude coordinates associated with a place.",
    "Agent":                         "Agent - An entity capable of acting, including persons, organizations, and groups.",
    "Place":                         "Place - A named geographic location associated with historical events or individuals.",
    "Person":                        "Person - An individual human being represented in historical records.",
    "ResearchProjectPIRole":         "ResearchProjectPIRole - The principal investigator role within a research project responsible for overseeing the work.",
    "ResearchProject":               "ResearchProject - A structured scholarly initiative aimed at collecting and analyzing historical data.",
    "Occupations":                   "Occupations - A set or list of occupations and skilled trades associated with individuals in historical records.",
    "ParticipantRoleTypes":          "ParticipantRoleTypes - A controlled vocabulary of roles that participants can hold in historical events.",
    "EntityWithProvenance":          "EntityWithProvenance - Any entity whose record includes documented provenance or source information.",
    "RaceRecord":                    "RaceRecord - A record capturing the racial classification assigned to an individual in a historical document.",
    "PersonStatusRecord":            "PersonStatusRecord - A record documenting the legal or social status of a person, such as enslaved, free, or manumitted.",
    "AgeCategory":                   "AgeCategory - A categorical grouping of age ranges used to classify individuals, such as child, adult, or elder.",
    "PlaceCV":                       "PlaceCV - A controlled vocabulary of place names used for geographic standardization.",
    "AgeRecord":                     "AgeRecord - A record capturing the age or estimated age of an individual at a specific point in time.",
    "PSCategories":                  "PSCategories - Categories of person status used to distinguish legal and social classifications of individuals.",
}


def classes_to_sentences(classes: list[str]) -> list[str]:
    """Return natural-language sentence for each class name."""
    return [CLASS_SENTENCES.get(c, f"An ontology class representing {c}.") for c in classes]


def split_camel(name: str) -> str:
    """CamelCase → space-separated words (fallback if no manual sentence)."""
    import re
    words = re.sub(r"([A-Z])", r" \1", name).strip().split()
    return " ".join(words)

# ──────────────────────────────────────────────────────────────────────────────
# MODEL REGISTRY
# ──────────────────────────────────────────────────────────────────────────────

MODEL_REGISTRY = {
    "bert": {
        "id":   "bert-base-uncased",
        "name": "BERT (bert-base-uncased)",
        "kwargs": {},
    },
    "bge": {
        "id":   "BAAI/bge-large-en-v1.5",
        "name": "BGE Large (BAAI/bge-large-en-v1.5)",
        "kwargs": {},
    },
    "qwen": {
        "id":   "Qwen/Qwen3-Embedding-8B",
        "name": "Qwen3 Embedding 8B",
        "kwargs": {"trust_remote_code": True},
    },
    "nvidia": {
        "id":   "nvidia/llama-embed-nemotron-8b",
        "name": "NVIDIA Nemotron 8B",
        "kwargs": {"trust_remote_code": True},
    },
}

# ──────────────────────────────────────────────────────────────────────────────
# EMBEDDING + SIMILARITY
# ──────────────────────────────────────────────────────────────────────────────

def load_model(model_key: str) -> SentenceTransformer:
    cfg = MODEL_REGISTRY[model_key]
    print(f"  Loading {cfg['name']} …")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    model = SentenceTransformer(cfg["id"], device=device, **cfg["kwargs"])
    return model


def compute_similarity(model: SentenceTransformer,
                        texts: list[str],
                        queries: list[str]) -> np.ndarray:
    """Return (n_queries × n_texts) cosine-similarity matrix."""
    print("    Encoding classes/sentences …")
    t_emb = model.encode(texts, convert_to_numpy=True, show_progress_bar=True,
                          normalize_embeddings=True)
    print("    Encoding competency questions …")
    q_emb = model.encode(queries, convert_to_numpy=True, show_progress_bar=True,
                          normalize_embeddings=True)
    return cosine_similarity(q_emb, t_emb)   # shape: (n_CQ, n_classes)

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

# Style constants
HDR_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_EVEN   = PatternFill("solid", start_color="DCE6F1")   # light blue
ROW_ODD    = PatternFill("solid", start_color="FFFFFF")
NUM_FONT   = Font(name="Arial", size=9)
CQ_FONT    = Font(name="Arial", size=9, bold=True)
TOP_BORDER = Border(top=Side(style="thin", color="1F4E79"))

# Gradient colours for similarity scores: white → green
def score_fill(score: float) -> PatternFill:
    """Map [0,1] → white-to-green hex fill."""
    r = int(255 * (1 - score))
    g = 200
    b = int(255 * (1 - score))
    hex_col = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill("solid", start_color=hex_col)

def write_sheet(ws, sim_matrix: np.ndarray, classes: list[str],
                queries: list[str], texts: list[str], to_sentence: bool):
    """Write one result sheet with Top-7 classes."""
    n_cq, n_cls = sim_matrix.shape
    TOP_K = 7

    # ── Header row ──
    headers = [
        "Competency Question",
        "Best Class",
        "Best Score",
        "Top 7 Classes",
        "Top 7 Scores"
    ]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Class headers
    for j, cls in enumerate(classes):
        label = texts[j] if to_sentence else cls
        cell = ws.cell(1, j + 6, label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # ── Data rows ──
    for i, cq in enumerate(queries):
        row = i + 2
        fill = ROW_EVEN if i % 2 == 0 else ROW_ODD

        # CQ
        c = ws.cell(row, 1, cq)
        c.font = CQ_FONT
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")

        scores = sim_matrix[i]

        # 🔥 Top-K logic (FIXED placement)
        top_indices = np.argsort(scores)[::-1][:TOP_K]
        top_classes = [classes[idx] for idx in top_indices]
        top_scores = [round(float(scores[idx]), 4) for idx in top_indices]

        # Best class
        bc = ws.cell(row, 2, top_classes[0])
        bc.font = Font(name="Arial", size=9, bold=True, color="1F4E79")
        bc.fill = fill
        bc.alignment = Alignment(horizontal="center", vertical="center")

        # Best score
        bs = ws.cell(row, 3, top_scores[0])
        bs.font = Font(name="Arial", size=9, bold=True)
        bs.fill = fill
        bs.number_format = "0.0000"
        bs.alignment = Alignment(horizontal="center", vertical="center")

        # Top 7 classes
        tc = ws.cell(row, 4, "\n".join(top_classes))
        tc.font = Font(name="Arial", size=9)
        tc.fill = fill
        tc.alignment = Alignment(wrap_text=True, vertical="center")

        # Top 7 scores
        ts = ws.cell(row, 5, "\n".join([f"{s:.4f}" for s in top_scores]))
        ts.font = Font(name="Arial", size=9)
        ts.fill = fill
        ts.alignment = Alignment(wrap_text=True, vertical="center")

        # All similarity scores
        for j, score in enumerate(scores):
            cell = ws.cell(row, j + 6, round(float(score), 4))
            cell.font = NUM_FONT
            cell.fill = score_fill(score)
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column widths ──
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18

    for j in range(n_cls):
        col_letter = get_column_letter(j + 6)
        ws.column_dimensions[col_letter].width = 22 if to_sentence else 24

    ws.row_dimensions[1].height = 55
    for i in range(n_cq):
        ws.row_dimensions[i + 2].height = 40

    ws.freeze_panes = "F2"


def write_summary_sheet(wb: Workbook, results: dict, classes: list[str], queries: list[str]):
    """Write a cross-model summary: for each CQ, which class ranks #1 per model."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "FF9900"

    headers = ["Competency Question"] + [r["model_name"] for r in results.values()]
    for j, h in enumerate(headers):
        c = ws.cell(1, j + 1, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for i, cq in enumerate(queries):
        row = i + 2
        fill = ROW_EVEN if i % 2 == 0 else ROW_ODD
        c = ws.cell(row, 1, cq)
        c.font = CQ_FONT
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")

        for j, (key, res) in enumerate(results.items()):
            best_idx = int(np.argmax(res["matrix"][i]))
            best_score = round(float(res["matrix"][i, best_idx]), 4)
            cell = ws.cell(row, j + 2, f"{classes[best_idx]}\n({best_score:.4f})")
            cell.font = Font(name="Arial", size=9)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 60
    for j in range(len(results)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 30
    ws.row_dimensions[1].height = 40
    for i in range(len(queries)):
        ws.row_dimensions[i + 2].height = 50
    ws.freeze_panes = "B2"

# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def run(model_keys: list[str], to_sentence: bool, output_path: str):
    texts = classes_to_sentences(CLASSES) if to_sentence else CLASSES
    mode_label = "sentence" if to_sentence else "raw"
    print(f"\n{'='*60}")
    print(f"Mode: {mode_label}  |  Models: {model_keys}")
    print(f"Classes: {len(CLASSES)}  |  CQs: {len(COMPETENCY_QUESTIONS)}")
    print(f"{'='*60}\n")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    results = {}
    for key in model_keys:
        cfg = MODEL_REGISTRY[key]
        print(f"\n[{key.upper()}] {cfg['name']}")
        t0 = time.time()
        model  = load_model(key)
        matrix = compute_similarity(model, texts, COMPETENCY_QUESTIONS)
        elapsed = time.time() - t0
        print(f"    Done in {elapsed:.1f}s")
        results[key] = {"matrix": matrix, "model_name": cfg["name"]}

        # One sheet per model
        sheet_name = f"{key.upper()}_{mode_label}"[:31]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = {
            "bert": "4472C4", "bge": "ED7D31", "qwen": "A9D18E", "nvidia": "FF0000"
        }.get(key, "808080")
        write_sheet(ws, matrix, CLASSES, COMPETENCY_QUESTIONS, texts, to_sentence)

    write_summary_sheet(wb, results, CLASSES, COMPETENCY_QUESTIONS)

    wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Enslaved ontology embedding evaluator")
    parser.add_argument("--to_sentence", action="store_true",
                        help="Convert class names to natural-language sentences before embedding (V2)")
    parser.add_argument("--both", action="store_true",
                        help="Run both raw and sentence modes and save to one workbook")
    parser.add_argument("--models", nargs="+", choices=list(MODEL_REGISTRY.keys()),
                        default=list(MODEL_REGISTRY.keys()),
                        help="Which models to run (default: all)")
    parser.add_argument("--output", default="embedding_results.xlsx",
                        help="Output Excel file path")
    args = parser.parse_args()

    if args.both:
        # Run both modes and write to a combined workbook
        wb = Workbook()
        wb.remove(wb.active)

        all_results_raw  = {}
        all_results_sent = {}

        for key in args.models:
            cfg = MODEL_REGISTRY[key]
            print(f"\n[{key.upper()}] {cfg['name']}")
            model = load_model(key)

            # --- raw ---
            matrix_raw = compute_similarity(model, CLASSES, COMPETENCY_QUESTIONS)
            all_results_raw[key] = {"matrix": matrix_raw, "model_name": cfg["name"] + " (raw)"}
            ws_raw = wb.create_sheet(f"{key.upper()}_raw"[:31])
            write_sheet(ws_raw, matrix_raw, CLASSES, COMPETENCY_QUESTIONS, CLASSES, False)

            # --- sentence ---
            texts_sent = classes_to_sentences(CLASSES)
            matrix_sent = compute_similarity(model, texts_sent, COMPETENCY_QUESTIONS)
            all_results_sent[key] = {"matrix": matrix_sent, "model_name": cfg["name"] + " (sent)"}
            ws_sent = wb.create_sheet(f"{key.upper()}_sent"[:31])
            write_sheet(ws_sent, matrix_sent, CLASSES, COMPETENCY_QUESTIONS, texts_sent, True)

        # Summary sheet for raw
        ws_sum_raw = wb.create_sheet("Summary_raw")
        ws_sum_raw.sheet_properties.tabColor = "4472C4"
        # reuse helper via temporary workbook trick
        tmp_raw = Workbook(); tmp_raw.remove(tmp_raw.active)
        write_summary_sheet(tmp_raw, all_results_raw, CLASSES, COMPETENCY_QUESTIONS)
        # copy rows
        src = tmp_raw["Summary"]
        dst = ws_sum_raw
        for row in src.iter_rows():
            for cell in row:
                nc = dst.cell(cell.row, cell.column, cell.value)
                if cell.font:   nc.font = cell.font.copy()
                if cell.fill:   nc.fill = cell.fill.copy()
                if cell.alignment: nc.alignment = cell.alignment.copy()
        for col, dim in src.column_dimensions.items():
            dst.column_dimensions[col].width = dim.width
        for row_num, dim in src.row_dimensions.items():
            dst.row_dimensions[row_num].height = dim.height
        dst.freeze_panes = "B2"

        # Summary sheet for sentence
        ws_sum_sent = wb.create_sheet("Summary_sent")
        ws_sum_sent.sheet_properties.tabColor = "ED7D31"
        tmp_sent = Workbook(); tmp_sent.remove(tmp_sent.active)
        write_summary_sheet(tmp_sent, all_results_sent, CLASSES, COMPETENCY_QUESTIONS)
        src2 = tmp_sent["Summary"]
        for row in src2.iter_rows():
            for cell in row:
                nc = ws_sum_sent.cell(cell.row, cell.column, cell.value)
                if cell.font:   nc.font = cell.font.copy()
                if cell.fill:   nc.fill = cell.fill.copy()
                if cell.alignment: nc.alignment = cell.alignment.copy()
        for col, dim in src2.column_dimensions.items():
            ws_sum_sent.column_dimensions[col].width = dim.width
        for row_num, dim in src2.row_dimensions.items():
            ws_sum_sent.row_dimensions[row_num].height = dim.height
        ws_sum_sent.freeze_panes = "B2"

        wb.save(args.output)
        print(f"\n✓ Saved combined workbook: {args.output}")
    else:
        run(args.models, args.to_sentence, args.output)


if __name__ == "__main__":
    main()